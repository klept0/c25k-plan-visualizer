"""
Core C25K logic module containing plan generation and export functions.
This module contains all the business logic that was previously in the main file.
"""

import json
import os
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional


def get_workout_details(week: int, day: int, lang: str = "e") -> str:
    """
    Get the detailed workout description for a specific week and day.
    Returns workout instructions based on the NHS C25K program.
    """
    # Complete C25K workout plan based on NHS guidelines
    workout_map = {
        1: {
            1: "Warm up: 5-min brisk walk. Run 60 sec, walk 90 sec (repeat 8 times). Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 60 sec, walk 90 sec (repeat 8 times). Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 60 sec, walk 90 sec (repeat 8 times). Cool down: 5-min walk.",
        },
        2: {
            1: "Warm up: 5-min brisk walk. Run 90 sec, walk 2 min (repeat 6 times). Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 90 sec, walk 2 min (repeat 6 times). Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 90 sec, walk 2 min (repeat 6 times). Cool down: 5-min walk.",
        },
        3: {
            1: "Warm up: 5-min brisk walk. Run 90 sec, walk 90 sec, run 3 min, walk 3 min (repeat 2 times). Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 90 sec, walk 90 sec, run 3 min, walk 3 min (repeat 2 times). Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 90 sec, walk 90 sec, run 3 min, walk 3 min (repeat 2 times). Cool down: 5-min walk.",
        },
        4: {
            1: "Warm up: 5-min brisk walk. Run 3 min, walk 90 sec, run 5 min, walk 2.5 min, run 3 min. Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 3 min, walk 90 sec, run 5 min, walk 2.5 min, run 3 min. Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 3 min, walk 90 sec, run 5 min, walk 2.5 min, run 3 min. Cool down: 5-min walk.",
        },
        5: {
            1: "Warm up: 5-min brisk walk. Run 5 min, walk 3 min, run 5 min, walk 3 min, run 5 min. Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 8 min, walk 5 min, run 8 min. Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 20 min (no walking breaks!). Cool down: 5-min walk.",
        },
        6: {
            1: "Warm up: 5-min brisk walk. Run 5 min, walk 3 min, run 8 min, walk 3 min, run 5 min. Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 10 min, walk 3 min, run 10 min. Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 25 min (no walking breaks!). Cool down: 5-min walk.",
        },
        7: {
            1: "Warm up: 5-min brisk walk. Run 25 min. Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 25 min. Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 25 min. Cool down: 5-min walk.",
        },
        8: {
            1: "Warm up: 5-min brisk walk. Run 28 min. Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 28 min. Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 28 min. Cool down: 5-min walk.",
        },
        9: {
            1: "Warm up: 5-min brisk walk. Run 30 min. Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 30 min. Cool down: 5-min walk.",
            3: "Warm up: 5-min brisk walk. Run 30 min. Cool down: 5-min walk.",
        },
        10: {
            1: "Warm up: 5-min brisk walk. Run 30+ min or 5K distance. Cool down: 5-min walk.",
            2: "Warm up: 5-min brisk walk. Run 30+ min or 5K distance. Cool down: 5-min walk.",
            3: "Graduation Day! Warm up: 5-min brisk walk. Run 30+ min or 5K distance. Cool down: 5-min walk. Congratulations, you did it!",
        },
    }

    default_workout = f"Week {week}, Day {day} C25K session - continue your training!"
    return workout_map.get(week, {}).get(day, default_workout)


def get_beginner_tip(day: int, lang: str = "e") -> str:
    """
    Get a beginner tip for the given day.
    """
    tips = [
        "Start slow and focus on endurance, not speed. Your pace should allow you to hold a conversation.",
        "Listen to your body and rest when needed. It's better to take an extra rest day than get injured.",
        "Stay hydrated before, during, and after running. Carry water on longer runs.",
        "Proper footwear is essential for injury prevention. Visit a running store for gait analysis.",
        "Consistency is more important than intensity. Aim to complete each session rather than go fast.",
        "Rest days are just as important as workout days. Your muscles grow and repair during rest.",
        "Celebrate small victories along the way! Every completed session is an achievement.",
        "Focus on your breathing. Try to breathe in through your nose and out through your mouth.",
        "Warm up properly with a brisk 5-minute walk before each running session.",
        "Cool down with gentle walking and light stretching to prevent stiffness.",
        "Consider running with a friend or joining a local C25K group for motivation.",
        "Track your progress in a journal or app to see how far you've come.",
        "Don't be discouraged if you need to repeat a week - everyone progresses at their own pace.",
        "Plan your route in advance and choose safe, well-lit paths.",
        "Listen to upbeat music or podcasts to keep yourself motivated during runs.",
        "Eat a light snack 30 minutes before running if you need energy, but avoid heavy meals.",
        "Focus on landing mid-foot rather than on your heels to reduce impact.",
        "Keep your arms relaxed and swing them naturally at your sides.",
        "If you feel pain (not just discomfort), stop and rest. Consult a doctor if pain persists.",
        "Remember why you started this journey and keep that motivation in mind during tough sessions.",
    ]

    return tips[day % len(tips)]


def get_weather_suggestion(location: str, session_date: datetime) -> str:
    """
    Get weather suggestions for a workout session.
    This is a simplified implementation - in a full version, you'd integrate with a weather API.
    """
    import random

    # Simulate different weather conditions
    conditions = [
        "Clear skies - perfect for running! Temperature around 65°F (18°C).",
        "Partly cloudy - good running weather. Light breeze expected.",
        "Overcast but dry - ideal running conditions, cooler temperatures.",
        "Light rain possible - consider indoor alternatives or waterproof gear.",
        "Hot and sunny - run early morning or evening, stay hydrated!",
        "Cool and crisp - great running weather, dress in layers.",
        "Windy conditions - choose a sheltered route if possible.",
    ]

    # For demo purposes, return a random condition based on the date
    random.seed(session_date.day + session_date.month)
    condition = random.choice(conditions)

    return f"Weather forecast for {location}: {condition}"


def get_workout_plan(user: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Return a plan (list of dicts) based on age, weight, and plan customization.
    Adjusts session duration for older or heavier users (age >= 60 or weight >= 100kg) for safety.
    Adds actual workout details and tips for each session.
    Supports custom rest day scheduling and adaptive plan logic.
    """
    weeks = user.get("weeks", 10)
    days_per_week = user.get("days_per_week", 3)
    rest_days = user.get("rest_days", ["Sat", "Sun"])
    plan: List[Dict[str, Any]] = []
    start_day = user.get("start_day")

    if not start_day:
        # Use today's date if not specified
        start_day = datetime.now()

    # Determine weight and threshold in correct units
    weight = user["weight"]
    weight_unit = user.get("weight_unit", "i")
    weight_str = f"{weight:.1f} lbs" if weight_unit == "i" else f"{weight:.1f} kg"
    weight_threshold = 220 if weight_unit == "i" else 100  # 220 lbs ≈ 100 kg

    for week in range(weeks):
        for day in range(days_per_week):
            day_offset = day * (7 // days_per_week)
            session_date = start_day + timedelta(weeks=week, days=day_offset)
            workout = get_workout_details(week + 1, day + 1, user.get("lang", "e"))
            tip = get_beginner_tip(day + 1, user.get("lang", "e"))

            # Fetch weather for this session
            weather_str = ""
            if user.get("location"):
                weather_str = get_weather_suggestion(user["location"], session_date)
            else:
                weather_str = "No location provided - check weather forecast before your workout!"

            description = (
                f"Follow the Couch to 5K plan - Week {week+1} session. "
                f"Note: This plan is tailored for an adult {user['gender']} "
                f"aged {user['age']} with hypertension. "
                f"Weight: {weight_str}. "
                f"Session time: {user['hour']:02d}:{user['minute']:02d}. "
                "Please monitor your health and consult your doctor if needed.\n"
                f"Workout: {workout}\n"
                f"Tip: {tip}"
            )

            plan.append(
                {
                    "week": week + 1,
                    "day": day + 1,
                    "day_offset": day_offset,
                    "duration": 30,  # minutes
                    "description": description,
                    "workout": workout,
                    "tip": tip,
                    "weather": weather_str,
                    "date": session_date.strftime("%Y-%m-%d"),
                }
            )

    # Adjust duration for older or heavier users
    if user["age"] >= 60 or weight >= weight_threshold:
        for session in plan:
            session["duration"] = 25
            session["description"] += " (Reduced session duration for safety.)"

    # Add rest days (user-selected)
    for week in range(weeks):
        for rest_offset, rest_name in enumerate(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]):
            if rest_name in rest_days:
                rest_date = start_day + timedelta(weeks=week, days=rest_offset)
                plan.append(
                    {
                        "week": week + 1,
                        "day": rest_name,
                        "day_offset": rest_offset,
                        "duration": 0,
                        "description": (
                            f"Rest Day - Week {week+1} {rest_name}. "
                            "Rest and recover. Hydrate and stretch."
                        ),
                        "workout": "Rest Day",
                        "tip": get_beginner_tip(rest_offset, user.get("lang", "e")),
                        "weather": "",
                        "date": rest_date.strftime("%Y-%m-%d"),
                    }
                )

    return sorted(plan, key=lambda s: (s["week"], s["day_offset"]))


def format_ics_datetime(dt: datetime) -> str:
    """
    Format a datetime object for ICS file (YYYYMMDDTHHMMSS).
    """
    return dt.strftime("%Y%m%dT%H%M%S")


def generate_ics(
    plan: List[Dict[str, Any]],
    start_day: datetime,
    hour: int,
    minute: int,
    alert_minutes: int = 30,
    outdir: str = ".",
) -> None:
    """
    Generate the ICS file from the workout plan and start date.
    Add the actual workout, tip, and rest days to the DESCRIPTION and NOTES fields.
    Add a VALARM block for custom alert time if alert_minutes > 0.
    ICS format is compatible with Apple/Google Calendar and most calendar apps.
    """
    ics_content = "BEGIN:VCALENDAR\nVERSION:2.0\nPRODID:-//Couch to 5K//EN\n"

    for session in plan:
        session_date = start_day + timedelta(weeks=session["week"] - 1, days=session["day_offset"])
        dt_start = session_date.replace(hour=hour, minute=minute)
        dt_end = dt_start + timedelta(minutes=session["duration"])

        if session["duration"] > 0:
            event_name = f"C25K Week {session['week']} - Day {session['day']}"
        else:
            event_name = f"C25K Week {session['week']} {session['day']} (Rest)"

        ics_content += (
            f"BEGIN:VEVENT\n"
            f"DTSTART:{format_ics_datetime(dt_start)}\n"
            f"DTEND:{format_ics_datetime(dt_end)}\n"
            f"SUMMARY:{event_name}\n"
            f"DESCRIPTION:{session['description']}\n"
            f"UID:{session['week']}-{session['day']}-c25k@couch-to-5k.local\n"
        )

        if alert_minutes > 0:
            ics_content += (
                f"BEGIN:VALARM\n"
                f"TRIGGER:-PT{alert_minutes}M\n"
                f"ACTION:DISPLAY\n"
                f"DESCRIPTION:C25K Reminder\n"
                f"END:VALARM\n"
            )

        ics_content += "END:VEVENT\n"

    ics_content += "END:VCALENDAR"

    # Save the ICS file
    filename = os.path.join(outdir, "c25k_plan.ics")
    with open(filename, "w", encoding="utf-8") as f:
        f.write(ics_content)


def generate_csv(plan: List[Dict[str, Any]], outdir: str = ".") -> None:
    """
    Generate a CSV file from the workout plan.
    """
    import csv

    filename = os.path.join(outdir, "c25k_plan.csv")

    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        fieldnames = ["week", "day", "date", "duration", "workout", "tip", "weather"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for session in plan:
            writer.writerow(
                {
                    "week": session["week"],
                    "day": session["day"],
                    "date": session["date"],
                    "duration": session["duration"],
                    "workout": session["workout"],
                    "tip": session["tip"],
                    "weather": session.get("weather", ""),
                }
            )


def generate_json(plan: List[Dict[str, Any]], outdir: str = ".") -> None:
    """
    Generate a JSON file from the workout plan.
    """
    filename = os.path.join(outdir, "c25k_plan.json")

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(plan, f, indent=2, ensure_ascii=False)


def generate_markdown(plan: List[Dict[str, Any]], user: Dict[str, Any], outdir: str = ".") -> None:
    """
    Generate a Markdown checklist from the workout plan.
    """
    filename = os.path.join(outdir, "c25k_checklist.md")

    content = f"""# C25K Training Plan

**Name:** {user.get('name', 'N/A')}
**Age:** {user.get('age', 'N/A')}
**Weight:** {user.get('weight', 'N/A')} {user.get('weight_unit', 'lbs')}
**Start Date:** {user.get('start_day', 'N/A')}

## Workout Schedule

"""

    current_week = 0
    for session in plan:
        if session["week"] != current_week:
            current_week = session["week"]
            content += f"\n### Week {current_week}\n\n"

        if session["duration"] > 0:
            content += f"- [ ] **Day {session['day']}** ({session['date']}): {session['workout']}\n"
            content += f"  - *Tip:* {session['tip']}\n"
        else:
            content += f"- [ ] **Rest Day** ({session['date']}): {session['workout']}\n"

    with open(filename, "w", encoding="utf-8") as f:
        f.write(content)


def generate_apple_health_csv(plan: List[Dict[str, Any]], outdir: str = ".") -> None:
    """
    Generate an Apple Health compatible CSV file from the workout plan.
    """
    import csv

    filename = os.path.join(outdir, "c25k_apple_health.csv")

    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        # Apple Health CSV format
        fieldnames = [
            "Start Date",
            "End Date",
            "Workout Type",
            "Duration (minutes)",
            "Distance (miles)",
            "Calories",
        ]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for session in plan:
            if session["duration"] > 0:  # Only workout days, not rest days
                start_time = f"{session['date']} 07:00:00"
                end_time = f"{session['date']} 07:{session['duration']:02d}:00"

                writer.writerow(
                    {
                        "Start Date": start_time,
                        "End Date": end_time,
                        "Workout Type": "Running",
                        "Duration (minutes)": session["duration"],
                        "Distance (miles)": "0.5-3.0",  # Estimated based on C25K progression
                        "Calories": int(session["duration"] * 10),  # Rough estimate
                    }
                )


def generate_strava_csv(plan: List[Dict[str, Any]], outdir: str = ".") -> None:
    """
    Generate a Strava-compatible CSV file from the workout plan.
    """
    import csv

    filename = os.path.join(outdir, "c25k_strava.csv")

    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        # Strava CSV format
        fieldnames = ["Activity Name", "Activity Type", "Date", "Time", "Duration", "Description"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for session in plan:
            if session["duration"] > 0:  # Only workout days
                writer.writerow(
                    {
                        "Activity Name": f"C25K Week {session['week']} Day {session['day']}",
                        "Activity Type": "Run",
                        "Date": session["date"],
                        "Time": "07:00:00",
                        "Duration": f"{session['duration']}:00",
                        "Description": session["workout"],
                    }
                )


def generate_excel_tracker(
    plan: List[Dict[str, Any]], user: Dict[str, Any], outdir: str = "."
) -> None:
    """
    Generate an Excel progress tracker with formulas and formatting.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        filename = os.path.join(outdir, "c25k_progress_tracker.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "C25K Progress"

        # Headers
        headers = [
            "Week",
            "Day",
            "Date",
            "Workout",
            "Completed",
            "Notes",
            "Effort (1-5)",
            "Weather",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, col=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = 2
        for session in plan:
            if session["duration"] > 0:  # Only workout days
                ws.cell(row=row, col=1, value=session["week"])
                ws.cell(row=row, col=2, value=session["day"])
                ws.cell(row=row, col=3, value=session["date"])
                ws.cell(row=row, col=4, value=session["workout"])

                # Completed checkbox (Y/N)
                completed_cell = ws.cell(row=row, col=5, value="N")
                completed_cell.alignment = Alignment(horizontal="center")

                # Notes
                ws.cell(row=row, col=6, value="")

                # Effort rating
                effort_cell = ws.cell(row=row, col=7, value="")
                effort_cell.alignment = Alignment(horizontal="center")

                # Weather
                ws.cell(row=row, col=8, value=session.get("weather", ""))

                row += 1

        # Add conditional formatting for completed sessions
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        ws.conditional_formatting.add(
            f"E2:E{row-1}", CellIsRule(operator="equal", formula=['"Y"'], fill=green_fill)
        )

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws["A1"] = "C25K Progress Summary"
        summary_ws["A1"].font = Font(size=16, bold=True)

        summary_ws["A3"] = "Total Sessions:"
        summary_ws["B3"] = f"=COUNTA('{ws.title}'!E:E)-1"

        summary_ws["A4"] = "Completed Sessions:"
        summary_ws["B4"] = f"=COUNTIF('{ws.title}'!E:E,\"Y\")"

        summary_ws["A5"] = "Completion Rate:"
        summary_ws["B5"] = f"=B4/B3*100"
        summary_ws["C5"] = "%"

        wb.save(filename)

    except ImportError:
        # Fallback to simple CSV if openpyxl not available
        generate_csv(plan, outdir)
